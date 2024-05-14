import os
from openpyxl import load_workbook

cwd = os.getcwd()
# 단어 엑셀 파일이 있는 경로이다.
voca_file_path = f"{cwd}\\main\\DB\\WordList.xlsx"

class VocaDBManager:

    # VocaFileManager는 파일의 경로를 매개변수로 가진다.
    def __init__(self, voca_file_path = f"{cwd}\\main\\DB\\WordList.xlsx"):
        self.voca_file_path = voca_file_path

        # Load the Excel workbook        
        self.wb = load_workbook(voca_file_path)

        # WordList.xlsx에 있는 네 개의 시트들이다.
        # 네 개의 시트들은 wordsheets 배열에 담긴다.
        self.ws1 = self.wb['wordsheet1']
        self.ws2 = self.wb['wordsheet2']
        self.ws3 = self.wb['wordsheet3']
        self.ws4 = self.wb['wordsheet4']

        self.wordsheets = [self.ws1, self.ws2, self.ws3, self.ws4]
        
    #단어가 존재하는지를 확인하는 함수
    def word_exists(self, word_name):
        wordsheets = self.wordsheets
        for ws in wordsheets:
            # B열을 탐색하여 해당 word_name이 존재하는지 탐색한다.
            for row in ws.iter_rows(min_row=2, min_col=1, max_col=2, max_row=ws.max_row, values_only=True):
                if word_name in row:
                    return True
        return False
    
    
    # 엑셀에서 해당 단어의 열 번호를 반환한다.
    def get_row_loc_of_word(self, word_name):
        if not (self.word_exists(word_name)):
            return False
        
        wordsheets = self.wordsheets
        for ws in wordsheets:
            idx_counter = 1
            for row in ws.iter_rows(min_row=2, min_col=1, max_col=2, max_row=ws.max_row, values_only=True):
                idx_counter += 1
                if word_name in row:
                    break
        return idx_counter 

    def delete_word(self,word_name):
        if(self.word_exists(word_name)):
            # 해당 영어 단어가 있는 행을 지워라.
            pass
        else:
            return False


    # 아직 구현 안 됨   
    # def edit_word(self, word_name, kor_meaning, word_class):
    #     if(self.word_exists(word_name)):
    #         wb = self.wb
    #         ws = self.ws
    #         row_loc = self.get_row_loc_of_word(word_name)
    #         if kor_meaning != "":
    #             ws[f'B{row_loc}'] = kor_meaning
    #         if word_class != "":
    #              ws[f'C{row_loc}'] = word_class
    #         wb.save(voca_file_path)
    #     else:
    #         return False
        
    
    # 단어를 추가하는 메소드이다.
    # word_class는 품사라는 뜻이다. 
    def add_word(self, word_name, kor_meaning, word_class, sheet_loc):
        
        # Check if the word is already in the Excel file
        if self.word_exists(word_name):
            print(f"The word '{word_name}' already exists in DB.")
            print("Go to Edit Word page.")
            return False

        if (not sheet_loc in (1,2,3,4)):
            print("sheet_loc 값이 잘못되었습니다.")
            return False

        ws = self.wordsheets[sheet_loc - 1]
         
        # Find the first empty row in the worksheet
        empty_row = ws.max_row + 1
        
        # Write the data to the next empty row
        ws[f'A{empty_row}'] = ws.max_row
        ws[f'B{empty_row}'] = word_name
        ws[f'C{empty_row}'] = kor_meaning
        ws[f'D{empty_row}'] = word_class
        
        # Save the changes to the Excel file
        self.wb.save(voca_file_path)

    def get_word_record(self, word_name):
        if not self.word_exists(word_name):
            return False
        
        wordsheets = self.wordsheets
        for ws in wordsheets:
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
                if word_name in row:
                    break
        return row
        
    

    # DB의 모든 단어의 array를 반환한다.
    def get_all_word_records(self):
        wordsheets = self.wordsheets
        words = []
        for ws in wordsheets:
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
                if (not row[0] == None):
                    words.append(row)
        return words
    
    def get_word_records_by_sheetloc(self, sheet_loc):
        if (not sheet_loc in (1,2,3,4)):
            print("sheet_loc 값이 잘못되었습니다.")
            return False
        ws = self.wordsheets[sheet_loc - 1]
        words = []
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
            if (not row[0] == None):
                words.append(row)
        return words

    # array를 return할 예정 
    def get_words_in_level(self, level=0):
        pass
    


voca_manager = VocaDBManager(voca_file_path)
#voca_manager.add_word("asdfasdfsdaf","뜻","v",3)

words = voca_manager.get_all_word_records()
print(words)



