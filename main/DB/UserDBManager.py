import os
from openpyxl import load_workbook

cwd = os.getcwd()

user_file_path = f"{cwd}\\main\\DB\\UserList.xlsx"

wb = load_workbook(user_file_path)

class UserDBManager:
    # VocaFileManager는 파일의 경로를 매개변수로 가진다.
    def __init__(self, user_file_path = f"{cwd}\\main\\DB\\UserList.xlsx"):
        self.voca_file_path = user_file_path
      
        self.wb = load_workbook(user_file_path)
        # usersheet
        self.ws = self.wb['usersheet']
        
    def get_user_record(self, user_id):
        if self.user_exists(user_id):
            ws = self.ws
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
                if user_id in row:
                    break
            # row는 다음과 같은 형태이다.
            # ('ad1', 'adpw1', '정준화', 'admin', 0)
            return row
        return False

    # 해당 유저가 있는 행의 번호를 반환한다.
    def get_idx_of_user(self, user_id):
        if self.word_exists(user_id):
            idx_counter = 1
            ws = self.ws
            for row in ws.iter_rows(min_row=2, max_col=1, max_row=ws.max_row, values_only=True):
                idx_counter += 1
                if user_id in row:
                    break
            return idx_counter
        return False

    def user_exists(self, user_id):
        ws = self.ws
        # Iterate through each row in column ID
        for row in ws.iter_rows(min_row=2, max_col=1, max_row=ws.max_row, values_only=True):
            if user_id in row:
                return True
        return False

    # 해당 유저의 패스워드를 가져온다.
    def get_user_password(self, user_id):
        if self.user_exists(user_id):
            user_password = self.get_user_record(user_id)[1]
            return user_password
        return False

    # 해당 유저의 레벨을 가져온다.
    def get_user_level(self, user_id):
        if self.user_exists(user_id):
            user_level = self.get_user_record(user_id)[4]
            return user_level
        return False

    # 전체 유저의 점수대별 분포를 보여준다.
    def get_dist():
        pass


um = UserDBManager()
print(um.get_user_record("ad1"))