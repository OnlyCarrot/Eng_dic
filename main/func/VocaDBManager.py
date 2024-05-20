from openpyxl import load_workbook
import os
import sys
sys.path.append(os.path.dirname(os.path.abspath(__file__)) + '/../../..')


# 테스트
"""
# some_file.py
import sys
# caution: path[0] is reserved for script path (or '' in REPL)
sys.path.insert(1, '/path/to/application/app/folder')

import file
"""

cwd = os.getcwd()
# 단어 엑셀 파일이 있는 경로이다.
voca_file_path = f"{cwd}/main/DB/WordList.xlsx"

class VocaDBManager:

    # VocaFileManager는 파일의 경로를 매개변수로 가진다.
    def __init__(self, voca_file_path = f"{cwd}/main/DB/WordList.xlsx"):
        self.voca_file_path = voca_file_path

        # Load the Excel workbook        
        self.wb = load_workbook(voca_file_path)

        # WordList.xlsx에 있는 네 개의 시트들이다.
        # 네 개의 시트들은 wordsheets 배열에 담긴다.
        self.ws1 = self.wb['wordsheet1']
        self.ws2 = self.wb['wordsheet2']
        self.ws3 = self.wb['wordsheet3']
        self.ws4 = self.wb['wordsheet4']

        self.wordsheets = [self.ws1, self.ws2, self.ws3, self.ws4]

        
    #단어가 존재하는지를 확인하는 함수
    def word_exists(self, word_name):
        wordsheets = self.wordsheets
        for ws in wordsheets:
            # B열을 탐색하여 해당 word_name이 존재하는지 탐색한다.
            for row in ws.iter_rows(min_row=2, max_col=1, max_row=ws.max_row, values_only=True):
                if word_name in row:
                    return True
        return False
    
    
    # 엑셀에서 해당 단어의 열 번호를 반환한다.
    def get_row_loc_of_word(self, word_name):
        """
        get_row_loc_of_word는 단어의 열 번호를 반환합니다.
        """

        if not (self.word_exists(word_name)):
            return False
        
        wordsheets = self.wordsheets
        for ws in wordsheets:
            idx_counter = 1
            for row in ws.iter_rows(min_row=2, max_col=1, max_row=ws.max_row, values_only=True):
                idx_counter += 1
                if word_name in row:
                    return idx_counter
        return False
        
    def delete_word(self,word_name):
        if(not self.word_exists(word_name)):
            return False
        sheet_num = self.get_level_of_word(word_name)
        ws = self.wordsheets[sheet_num-1]
        row_idx = self.get_row_loc_of_word(word_name)
        ws[f'A{row_idx}'] = ""
        ws[f'B{row_idx}'] = ""
        ws[f'C{row_idx}'] = ""
        self.wb.save(voca_file_path)
        return;

    #아직 구현 안 됨   
    def edit_word(self, word_name, kor_meaning, word_class):
        if(not self.word_exists(word_name)):
            return False
        
        sheet_num = self.get_level_of_word(word_name)
        row_idx = self.get_row_loc_of_word(word_name)
        ws = self.wordsheets[sheet_num - 1]
        ws[f'B{row_idx}'] = kor_meaning
        ws[f'C{row_idx}'] = word_class
        self.wb.save(voca_file_path)
        return    
    
    # 단어를 추가하는 메소드이다.
    # word_class는 품사라는 뜻이다. 
    def add_word(self, word_name, kor_meaning, word_class, sheet_loc):
        
        # Check if the word is already in the Excel file
        if self.word_exists(word_name):
            print(f"The word '{word_name}' already exists in DB.")
            print("Go to Edit Word page.")
            return False

        if (not sheet_loc in (1,2,3,4)):
            print("sheet_loc 값이 잘못되었습니다.")
            return False

        ws = self.wordsheets[sheet_loc - 1]
         
        # Find the first empty row in the worksheet
        empty_row = ws.max_row + 1
        
        # Write the data to the next empty row
        ws[f'A{empty_row}'] = word_name
        ws[f'B{empty_row}'] = kor_meaning
        ws[f'C{empty_row}'] = word_class
        
        # Save the changes to the Excel file
        self.wb.save(voca_file_path)

    def get_word_record(self, word_name):
        if not self.word_exists(word_name):
            return False
        
        wordsheets = self.wordsheets
        for ws in wordsheets:
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
                if word_name in row:
                    return row
        return False
    
    def get_level_of_word(self, word_name):
        if not self.word_exists(word_name):
            return False
        
        wordsheets = self.wordsheets
        counter = 1
        for ws in wordsheets:
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
                if word_name in row:
                    return counter
            counter += 1
        return False
        

    # DB의 모든 단어의 array를 반환한다.
    def get_all_word_records(self):
        wordsheets = self.wordsheets
        words = []
        for ws in wordsheets:
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
                if (not row[0] == None):
                    words.append(row)
        return words
    
    def get_word_records_by_sheetloc(self, sheet_loc):
        if (not sheet_loc in (1,2,3,4)):
            print("sheet_loc 값이 잘못되었습니다.")
            return False
        ws = self.wordsheets[sheet_loc - 1]
        words = []
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
            if (not row[0] == None):
                words.append(row)
        return words

    # array를 return할 예정 
    def get_words_in_level(self, level=0):
        pass
    
    def sort_words(self, sheet_num):
        pass
        """
        sort_words는 단어를 영어 로마자 순서에 맞게 오름차순으로 정렬합니다.
        """


# words = voca_manager.get_word_records_by_sheetloc(3)
# print(words)



