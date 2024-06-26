from openpyxl import load_workbook
import os
import sys
sys.path.append(os.path.dirname(os.path.abspath(__file__)) + '/../../..')
from main.func.Sheet import Sheet
from tkinter import END

class Voca:
    def __init__(self):
        sheets = [
            Sheet("wordsheet1").worksheet,
            Sheet("wordsheet2").worksheet,
            Sheet("wordsheet3").worksheet,
            Sheet("wordsheet4").worksheet
        ]
        # 모든 시트 객체를 리스트로 저장
        self.sheets = sheets
            
    def show_voca(self, listbox): 

        # 리스트박스 초기화
        listbox.delete(0, END)

        i = 1
        for sheet in self.sheets:
            for row in sheet.iter_rows(2, values_only=True):
                eng, kor, c = row
                listbox.insert(END, f"{i}    {eng} - {kor} - {c}")
                i += 1
            
    def search_voca(self, input):
        find_word_list = []
        for sheet in self.sheets:
            for row in sheet.iter_rows(values_only=True):
                eng, kor, c = row
                if eng.startswith(input):
                    #print(eng, kor, c)
                    find_word_list.append([eng, kor, c])
        if(len(find_word_list) > 0):
            return find_word_list
        else:
            return False
