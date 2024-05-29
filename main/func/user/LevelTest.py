from openpyxl import load_workbook
import random
import os
import sys
sys.path.append(os.path.dirname(os.path.abspath(__file__)) + '/../../..')
from main.func.Sheet import Sheet

class LevelTest:
    
    total_score = 0
    word = []

    def select_word():
    # 이미 추출된 단어를 추적하기 위한 세트
        extracted_words = set()

        # 각 시트에서 10개씩 단어 추출
        for sheet_num in range(1, 5):
            sheet_name = f"wordsheet{sheet_num}"
            sheet = Sheet(sheet_name).worksheet

            count = 0
            # 각 시트에서 10번씩 랜덤한 단어 추출
            while count < 10:
                # 랜덤한 행 번호 생성
                random_row_num = random.randint(2, sheet.max_row)

                # 랜덤한 행의 데이터 추출
                for row in sheet.iter_rows(min_row=random_row_num, max_row=random_row_num, values_only=True):
                    word = (row[0], row[1])
                    if word not in extracted_words:
                        LevelTest.word.append([row[0], row[1]])
                        extracted_words.add(word)
                        count += 1
                        break

        return LevelTest.word


    def show_word_meaning1(self):
        wordlist1 = []
        # 단어 뜻을 보여줌 x좌표가 610 에 1~5번 단어
        for i in range(0, 5):
            self.canvas.create_text(460.0, 130.0 + i * 53.5, text=LevelTest.select_word()[i][1])
            wordlist1.append([LevelTest.select_word()[i][0], LevelTest.select_word()[i][1]])
        # 단어 뜻을 보여줌 x좌표가 950 에 6~10번 단어
        for i in range(0, 5):
            self.canvas.create_text(800.0, 130.0 + i * 53.5, text=LevelTest.select_word()[i + 5][1])
            wordlist1.append([LevelTest.select_word()[i + 5][0], LevelTest.select_word()[i + 5][1]])
        return wordlist1
    
    def show_word_meaning2(self):
        wordlist2 = []
        # 단어 뜻을 보여줌 x좌표가 610 에 11~15번 단어
        for i in range(0, 5):
            self.canvas.create_text(460.0, 130.0 + i * 53.5, text=LevelTest.select_word()[i + 10][1])
            wordlist2.append([LevelTest.select_word()[i + 10][0], LevelTest.select_word()[i + 10][1]])
        # 단어 뜻을 보여줌 x좌표가 950 에 16~20번 단어
        for i in range(0, 5):
            self.canvas.create_text(800.0, 130.0 + i * 53.5, text=LevelTest.select_word()[i + 15][1])
            wordlist2.append([LevelTest.select_word()[i + 15][0], LevelTest.select_word()[i + 15][1]])
        return wordlist2
    
    def show_word_meaning3(self):
        wordlist3 = []
        # 단어 뜻을 보여줌 x좌표가 610 에 21~25번 단어
        for i in range(0, 5):
            self.canvas.create_text(460.0, 130.0 + i * 53.5, text=LevelTest.select_word()[i + 20][1])
            wordlist3.append([LevelTest.select_word()[i + 20][0], LevelTest.select_word()[i + 20][1]])
        # 단어 뜻을 보여줌 x좌표가 950 에 26~30번 단어
        for i in range(0, 5):
            self.canvas.create_text(800.0, 130.0 + i * 53.5, text=LevelTest.select_word()[i + 25][1])
            wordlist3.append([LevelTest.select_word()[i + 25][0], LevelTest.select_word()[i + 25][1]])

        return wordlist3
    
    def show_word_meaning4(self):
        wordlist4 = []
        # 단어 뜻을 보여줌 x좌표가 610 에 31~35번 단어
        for i in range(0, 5):
            self.canvas.create_text(460.0, 130.0 + i * 53.5, text=LevelTest.select_word()[i + 30][1])
            wordlist4.append([LevelTest.select_word()[i + 30][0], LevelTest.select_word()[i + 30][1]])
        # 단어 뜻을 보여줌 x좌표가 950 에 36~40번 단어
        for i in range(0, 5):
            self.canvas.create_text(800.0, 130.0 + i * 53.5, text=LevelTest.select_word()[i + 35][1])
            wordlist4.append([LevelTest.select_word()[i + 35][0], LevelTest.select_word()[i + 35][1]])
        return wordlist4
    
    # user_input_word는 list로 받을거라 나중에 user_input_word[i] 로 수정
    def grade_score(user_input_word, word):
        score = 0
        for i in range(0, len(user_input_word)):
            correct_answer = word[i][0]
            if user_input_word[i] == correct_answer:
                score += 1

        return score
    
    def is_str_vaild(ans):
        if(ans.strip() ==''):
            print("공백 값이 입력되었습니다. 유효한 단어를 입력하세요.")
            return False
        if(not ans.replace(' ', '').isalpha()):
            print("알파벳이 아닌 문자열이 입력되었습니다.")
            return False
        return True
    
    def temp_score(score):
        LevelTest.total_score += score
        return LevelTest.total_score
    
    def set_user_level(final_score):
        if final_score <= 10:
            return 600
        elif final_score <= 20:
            return 700
        elif final_score <= 30:
            return 800
        else:
            return 900