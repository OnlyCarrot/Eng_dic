import os
import openpyxl
import sys
sys.path.append(os.path.dirname(os.path.abspath(__file__)) + '/../../..')
from func.Sheet import Sheet
from .AdminSearchVoca import word_exists, get_row_loc_of_word, get_level_of_word

# 엑셀 파일을 로드합니다.
file_path = r"C:\Eng_dic\interface\OnlyCarrot\WordList.xlsx"
wb = openpyxl.load_workbook(file_path)
sheet = wb['wordsheet']

# Tkinter 윈도우를 생성합니다.
root = tk.Tk()
root.title("영어 단어 검색")

# 오류 메시지 변수
error_message = None  # 오류 메시지를 저장하는 변수

# 검색 함수를 정의합니다.
def search_word():
    search_term = entry.get()   # 입력 상자에서 검색어를 가져옵니다.

    # 사용자가 아무것도 입력하지 않은 경우
    if not search_term:
        messagebox.showwarning("경고", "검색어를 입력하세요.")
        entry.delete(0, tk.END)  # 입력 필드를 지웁니다.
        return

    search_term = search_term.lower()   # 대소문자 구분 없이 검색하기 위해 소문자로 변환합니다.

    # 검색어를 포함하는 모든 단어를 찾습니다
    try:
        matching_words = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if search_term in row[0].lower():
                matching_words.append(row)

        if not matching_words:
            raise Exception("검색 결과가 없습니다.")

    except Exception as e:
        error_message = str(e)  # 오류 메시지를 저장합니다.

    else:
        show_results(matching_words)

def show_results(words):
    result_window = tk.Toplevel(root)
    result_window.title("검색 결과")

    listbox = tk.Listbox(result_window, width=50, height=10, font=("arial", 10))
    listbox.pack(padx=10, pady=10)

    for word in words:
        word_info = f"\n\t{word[0]} \u2014 {word[1]} ({word[2]})"  # 단어를 굵게 표시합니다.
        listbox.insert(tk.END, word_info)


def add_word(word_name, kor_meaning, word_class, word_level):
    """
        DB에 새로운 단어를 추가합니다.
        word_name: 영어 단어명
        kor_meaning: 한국어 해석
        word_class: 품사
        word_level: 단어의 레벨(1,2,3,4 중 하나)

    """
    sheet = Sheet()
    wordsheets = sheet.wordsheets

    # Check if the word is already in the Excel file
    if word_exists(word_name):
        print(f"The word '{word_name}' already exists in DB.")
        print("Go to Edit Word page.")
        return False

    if (not word_level in (1,2,3,4)):
        print("word_level 값이 잘못되었습니다.")
        return False
    
    ws = wordsheets[word_level - 1]

    # Find the first empty row in the worksheet
    empty_row = ws.max_row + 1
    
    # Write the data to the next empty row
    ws[f'A{empty_row}'] = word_name
    ws[f'B{empty_row}'] = kor_meaning
    ws[f'C{empty_row}'] = word_class
    
    # Save the changes to the Excel file
    sheet.save()

# GUI 요소를 생성합니다.
label = tk.Label(root, text="검색어를 입력하세요:")
label.pack(pady=5)

entry = tk.Entry(root, width=30)
entry.pack(pady=5)

search_button = tk.Button(root, text="검색", command=search_word)
search_button.pack(pady=5)

# GUI를 실행합니다.
root.mainloop()

# 결과 처리
if error_message:
    messagebox.showinfo("검색 결과", error_message)