from openpyxl import load_workbook
import os
import sys
sys.path.append(os.path.dirname(os.path.abspath(__file__)) + '/../../..')
from func.Sheet import Sheet

class UserInfo:
    def __init__(self):
        self.sheet = Sheet("usersheet").worksheet
        
    def show_selected_user(self, user_id):
        for row in self.sheet.iter_rows(values_only=True):
            id, password, name, role, level, last_date = row
            if(user_id == id):
                print(id, password, name, role, level, last_date)

    def show_all_user(self):
        for row in self.sheet.iter_rows(values_only=True):
            id, password, name, role, level, last_date = row
            if(not id.startswith("ad")):
                print(id, password, name, role, level, last_date)

    def show_total_user_count(self):
        total_user_count = 0 
        for row in self.sheet.iter_rows(values_only=True):
            for role in row:
                if role == "user":
                    total_user_count += 1
        print("총 사용자 수:", total_user_count)
        


    # 숫자 표기를 통한 분포도 확인
    def show_distribution_user(self):
        levels_count = {
            "~700": 0,
            "700+": 0,
            "800+": 0,
            "900+": 0
        }

        for row in self.sheet.iter_rows(values_only=True):
            level = row[-2]

            try:
                level = int(level)
            except ValueError:
                # If conversion fails, skip this row
                continue
            
            # Determine the range for the level
            if level < 700:
                key = "~700"
            elif 700 <= level < 800:
                key = "700+"
            elif 800 <= level < 900:
                key = "800+"
            elif 900 <= level < 990:
                key = "900+"
            
            # Update the levels_count dictionary
            if key in levels_count:
                levels_count[key] += 1
            else:
                levels_count[key] = 1

        order = ["~700", "700+", "800+", "900+"]
        for level in order:
            count = levels_count.get(level, 0)
            print(f"{level}: {count}명")


def is_str_valid(ans):
    if ans.strip() == '':
        return False
    return True 