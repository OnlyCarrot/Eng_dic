from openpyxl import load_workbook
import os
cwd = os.getcwd()

# 단어 엑셀 파일이 있는 경로이다.
voca_file_path = f"{cwd}/main/DB/WordList.xlsx"

class Sheet:
    def __init__(self, sheet):
        if(sheet == "usersheet"):
            self.workbook = load_workbook("main/DB/UserList.xlsx")
            self.worksheet = self.workbook[sheet]
        else:
            self.workbook = load_workbook("main/DB/WordList.xlsx")
            self.worksheet = self.workbook[sheet]

            self.__ws1 = self.workbook['wordsheet1']
            self.__ws2 = self.workbook['wordsheet2']
            self.__ws3 = self.workbook['wordsheet3']
            self.__ws4 = self.workbook['wordsheet4']
            self.wordsheets = [self.__ws1, self.__ws2, self.__ws3, self.__ws4]

    def workbook(self):
        return self.workbook
    
    def worksheet(self):
        return self.worksheet
    
    def save(self):
        self.workbook.save(voca_file_path)
    

    

