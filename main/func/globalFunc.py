from openpyxl import load_workbook

class open_sheet:
    def wordsheet(wordsheet):
        # 엑셀 파일 열기
        workbook = load_workbook("main/DB/WordList.xlsx")

        # 시트 선택
        return workbook[wordsheet]

    def usersheet():
        # 엑셀 파일 열기
        workbook = load_workbook("main/DB/UserList.xlsx")

        # 시트 선택
        return workbook["usersheet"]

from openpyxl import load_workbook
import random
import os
import sys
sys.path.append(os.path.dirname(os.path.abspath(__file__)) + '/../../../..')
from func.globalFunc import open_sheet