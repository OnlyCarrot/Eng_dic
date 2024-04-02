import openpyxl as op

wb = op.load_workbook("C:\\Eng_dic\\Trash\\test_excel_py\\pytest.xlsx")

def pickup(sheet_name, search_value):
    ws = wb[sheet_name]
    for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if search_value in row:
            # search_value가 있는 열 인덱스
            col_index = row.index(search_value) + 1
            # idx는 이미 현재 행 번호를 나타내므로, 다음 행의 셀 값을 얻습니다.
            next_row_value = ws.cell(row=idx, column=col_index+1).value
            print(next_row_value)
            return  # 값을 찾았으니 함수 종료
    print("Value not found.")


# 함수 호출 예시
pickup("과일", "코코넛")
pickup("사물", "병원")
