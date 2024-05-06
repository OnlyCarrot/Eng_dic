import tkinter as tk
from tkinter import messagebox
import openpyxl

class AdminEdit(tk.Frame):
    def __init__(self, master):
        super().__init__(master)
        self.master = master

        label_title = tk.Label(self, text="단어 수정 (관리자 전용)", font=("Helvetica", 16, "bold"))
        label_title.pack(pady=20)

        self.search_entry = tk.Entry(self, width=30)
        self.search_entry.pack(pady=10)

        search_button = tk.Button(self, text="검색", command=self.search_word)
        search_button.pack(pady=10)

        self.meaning_entry = tk.Entry(self, width=30)
        self.meaning_entry.pack(pady=10)

        self.part_of_speech_entry = tk.Entry(self, width=30)
        self.part_of_speech_entry.pack(pady=10)

        update_button = tk.Button(self, text="수정", command=self.update_word)
        update_button.pack(pady=10)

    def search_word(self):
        """입력된 단어로 단어를 검색하여 뜻과 품사를 표시하는 함수"""
        search_word = self.search_entry.get().strip().lower()

        try:
            wb = openpyxl.load_workbook('word_data.xlsx')
            sheet = wb.active

            found = False

            for row in sheet.iter_rows(values_only=True):
                if row[0].lower() == search_word:
                    self.meaning_entry.delete(0, tk.END)
                    self.meaning_entry.insert(0, row[1])
                    self.part_of_speech_entry.delete(0, tk.END)
                    self.part_of_speech_entry.insert(0, row[2])
                    found = True
                    break

            if not found:
                messagebox.showwarning("검색 오류", "해당 단어를 찾을 수 없습니다.")
                self.meaning_entry.delete(0, tk.END)
                self.part_of_speech_entry.delete(0, tk.END)

        except FileNotFoundError:
            messagebox.showerror("파일 오류", "단어 데이터 파일을 찾을 수 없습니다.")

    def update_word(self):
        """검색된 단어의 뜻과 품사를 업데이트하는 함수"""
        search_word = self.search_entry.get().strip().lower()
        new_meaning = self.meaning_entry.get().strip()
        new_part_of_speech = self.part_of_speech_entry.get().strip()

        try:
            wb = openpyxl.load_workbook('word_data.xlsx')
            sheet = wb.active

            found = False

            for row in sheet.iter_rows():
                if row[0].value.lower() == search_word:
                    row[1].value = new_meaning
                    row[2].value = new_part_of_speech
                    found = True
                    break

            if found:
                wb.save('word_data.xlsx')
                messagebox.showinfo("수정 완료", "단어 정보가 수정되었습니다.")
            else:
                messagebox.showwarning("수정 오류", "해당 단어를 찾을 수 없습니다.")

        except FileNotFoundError:
            messagebox.showerror("파일 오류", "단어 데이터 파일을 찾을 수 없습니다.")

# 테스트 코드
if __name__ == "__main__":
    root = tk.Tk()
    root.title("단어 수정")

    admin_edit_frame = AdminEdit(root)
    admin_edit_frame.pack(fill=tk.BOTH, expand=True)

    root.mainloop()
