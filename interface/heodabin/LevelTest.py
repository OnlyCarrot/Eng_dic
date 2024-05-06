import tkinter as tk
from tkinter import messagebox
import openpyxl
import random

class LevelTest(tk.Tk):
    def __init__(self, user_id):
        """
        Attributes:
            user_id (str): 사용자의 아이디
            words (list): 엑셀 파일에서 읽어온 영어 단어 리스트
            meanings (list): 엑셀 파일에서 읽어온 한글 뜻 리스트
            current_question (int): 현재 테스트 중인 문제의 인덱스
            correct_count (int): 정답 수
            total_questions (int): 총 문제 수
        """
        super().__init__()
        self.title("Level Test")
        self.user_id = user_id
        self.words = []
        self.meanings = []
        self.current_question = 0
        self.correct_count = 0
        self.total_questions = 10  # 테스트할 문제 수 (임시)

        self.load_word_data()
        self.create_widgets()

    def load_word_data(self):
        """엑셀 파일에서 단어 데이터를 읽어와서 리스트에 저장하는 함수."""
        try:
            wb = openpyxl.load_workbook('word_data.xlsx')
            sheet = wb.active

            for row in sheet.iter_rows(values_only=True):
                self.words.append(row[0])
                self.meanings.append(row[1])

        except FileNotFoundError:
            messagebox.showerror("파일 오류", "단어 데이터 파일을 찾을 수 없습니다.")

    def create_widgets(self):
        """GUI 위젯을 생성하고 배치하는 함수."""
        self.label_question = tk.Label(self, text="한글 뜻: ")
        self.label_question.pack(pady=20)

        self.entry_answer = tk.Entry(self)
        self.entry_answer.pack(pady=10)

        self.button_submit = tk.Button(self, text="제출", command=self.check_answer)
        self.button_submit.pack(pady=10)

        self.update_question()

    def update_question(self):
        """다음 문제로 넘어가거나 테스트 종료를 처리하는 함수."""
        if self.current_question < self.total_questions:
            self.label_question.config(text=f"한글 뜻: {self.meanings[self.current_question]}")
        else:
            self.finish_test()

    def check_answer(self):
        """사용자의 답변을 검사하고 정답 여부를 판단하는 함수."""
        user_answer = self.entry_answer.get().strip().lower()
        correct_answer = self.words[self.current_question].lower()

        if user_answer == correct_answer:
            self.correct_count += 1

        self.current_question += 1
        self.entry_answer.delete(0, tk.END)
        self.update_question()

    def finish_test(self):
        """테스트 완료 후 정확도에 따라 레벨을 결정하고 저장하는 함수."""
        accuracy = self.correct_count / self.total_questions
        if accuracy >= 0.8:
            level = 3
        elif accuracy >= 0.6:
            level = 2
        else:
            level = 1

        messagebox.showinfo("테스트 완료", f"테스트가 완료되었습니다. 레벨: {level}")

        self.save_user_level(level)
        self.destroy()  # 테스트 창 닫기
        

    def save_user_level(self, level):
    #사용자의 레벨을 엑셀 파일에 저장하는 함수.
        try:
            wb = openpyxl.load_workbook('user_data.xlsx')
            sheet = wb.active

            for row in sheet.iter_rows():
                if row[1].value == self.user_id:
                    row[3].value = level  # 레벨 업데이트
                    break


            wb.save('user_data.xlsx')
            messagebox.showinfo("저장 완료", "사용자 레벨이 성공적으로 저장되었습니다.")

        except FileNotFoundError:
            messagebox.showerror("파일 오류", "사용자 데이터 파일을 찾을 수 없습니다.")
        except Exception as e:
            messagebox.showerror("오류", f"사용자 레벨 저장 오류: {e}")




