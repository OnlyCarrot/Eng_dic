import tkinter as tk
from tkinter import messagebox
import openpyxl

class AdminAdd(tk.Frame):
    def __init__(self, master):
        super().__init__(master)
        self.master = master

        label_title = tk.Label(self, text="단어 추가 (관리자 전용)", font=("Helvetica", 16, "bold"))
        label_title.pack(pady=20)

        self.word_entry = tk.Entry(self, width=30)
        self.word_entry.pack(pady=10)

        self.meaning_entry = tk.Entry(self, width=30)
        self.meaning_entry.pack(pady=10)

        add_button = tk.Button(self, text="단어 추가", command=self.add_word)
        add_button.pack(pady=10)

    def add_word(self):
        """입력된 단어를 word_data에 추가하는 함수"""
        new_word = self.word_entry.get().strip().lower()
        new_meaning = self.meaning_entry.get().strip()

        if not new_word or not new_meaning:
            messagebox.showwarning("입력 오류", "단어와 뜻을 모두 입력하세요.")
            return

        if self.check_duplicate(new_word):
            messagebox.showwarning("중복 오류", "이미 존재하는 단어입니다.")
            return

        try:
            wb = openpyxl.load_workbook('word_data.xlsx')
            sheet = wb.active

            sheet.append([new_word, new_meaning])
            wb.save('word_data.xlsx')

            messagebox.showinfo("단어 추가", "새로운 단어가 추가되었습니다.")
            self.word_entry.delete(0, tk.END)
            self.meaning_entry.delete(0, tk.END)

        except FileNotFoundError:
            messagebox.showerror("파일 오류", "단어 데이터 파일을 찾을 수 없습니다.")

    def check_duplicate(self, new_word):
        """새로 추가하려는 단어가 이미 존재하는지 확인하는 함수"""
        try:
            wb = openpyxl.load_workbook('word_data.xlsx')
            sheet = wb.active

            for row in sheet.iter_rows(values_only=True):
                if row[0].lower() == new_word:
                    return True

            return False

        except FileNotFoundError:
            return False

# 테스트 코드
if __name__ == "__main__":
    root = tk.Tk()
    root.title("단어 추가")

    admin_add_frame = AdminAdd(root)
    admin_add_frame.pack(fill=tk.BOTH, expand=True)

    root.mainloop()
