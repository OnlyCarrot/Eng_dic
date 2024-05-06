import tkinter as tk
from tkinter import messagebox
import openpyxl

class AdminDel(tk.Frame):
    def __init__(self, master):
        super().__init__(master)
        self.master = master

        label_title = tk.Label(self, text="단어 삭제 (관리자 전용)", font=("Helvetica", 16, "bold"))
        label_title.pack(pady=20)

        self.search_entry = tk.Entry(self, width=30)
        self.search_entry.pack(pady=10)

        search_button = tk.Button(self, text="검색", command=self.search_word)
        search_button.pack(pady=10)

        delete_button = tk.Button(self, text="삭제", command=self.delete_word)
        delete_button.pack(pady=10)

    def search_word(self):
        """입력된 단어로 단어를 검색하여 존재 여부를 확인하는 함수"""
        search_word = self.search_entry.get().strip().lower()

        try:
            wb = openpyxl.load_workbook('word_data.xlsx')
            sheet = wb.active

            found = False

            for row in sheet.iter_rows():
                if row[0].value.lower() == search_word:
                    found = True
                    break

            if found:
                messagebox.showinfo("검색 결과", f"'{search_word}' 단어가 존재합니다.")
            else:
                messagebox.showwarning("검색 결과", f"'{search_word}' 단어를 찾을 수 없습니다.")

        except FileNotFoundError:
            messagebox.showerror("파일 오류", "단어 데이터 파일을 찾을 수 없습니다.")

    def delete_word(self):
        """검색된 단어를 삭제하는 함수"""
        search_word = self.search_entry.get().strip().lower()

        try:
            wb = openpyxl.load_workbook('word_data.xlsx')
            sheet = wb.active

            found = False

            for row in sheet.iter_rows():
                if row[0].value.lower() == search_word:
                    sheet.delete_rows(row[0].row)
                    found = True
                    break

            if found:
                wb.save('word_data.xlsx')
                messagebox.showinfo("삭제 완료", f"'{search_word}' 단어가 삭제되었습니다.")
                self.search_entry.delete(0, tk.END)
            else:
                messagebox.showwarning("삭제 오류", f"'{search_word}' 단어를 찾을 수 없습니다.")
                self.search_entry.delete(0, tk.END)

        except FileNotFoundError:
            messagebox.showerror("파일 오류", "단어 데이터 파일을 찾을 수 없습니다.")

# 테스트 코드
if __name__ == "__main__":
    root = tk.Tk()
    root.title("단어 삭제")

    admin_del_frame = AdminDel(root)
    admin_del_frame.pack(fill=tk.BOTH, expand=True)

    root.mainloop()
