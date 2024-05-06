import tkinter as tk
from openpyxl import load_workbook

def read_excel(file_path):
    """
    엑셀 파일을 읽어와서 데이터를 반환하는 함수
    """
    wb = load_workbook(file_path)
    ws = wb['wordsheet']

    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        eng_word, kor_meaning, word_class = row
        data.append((eng_word, kor_meaning, word_class))

    return data
def show_words(word_data, start_index):
    """
    리스트박스에 단어를 보여주는 함수
    """
    window = tk.Tk()
    window.title("영어 단어 조회")

    listbox = tk.Listbox(window, width=50, height=10)
    listbox.grid(row=0, column=0, padx=10, pady=10, columnspan=2)

    # 단어 데이터를 리스트박스에 추가
    for i in range(start_index, min(start_index + 10, len(word_data))):
        word, meaning, word_class = word_data[i]
        listbox.insert(tk.END, f"{word}: {meaning} ({word_class})")

    def next_words():
        nonlocal start_index
        start_index += 10
        listbox.delete(0, tk.END)
        for i in range(start_index, min(start_index + 10, len(word_data))):
            word, meaning, word_class = word_data[i]
            listbox.insert(tk.END, f"{word}: {meaning} ({word_class})")

    def prev_words():
        nonlocal start_index
        start_index -= 10
        if start_index < 0:
            start_index = 0
        listbox.delete(0, tk.END)
        for i in range(start_index, min(start_index + 10, len(word_data))):
            word, meaning, word_class = word_data[i]
            listbox.insert(tk.END, f"{word}: {meaning} ({word_class})")

    next_button = tk.Button(window, text="다음", command=next_words)
    next_button.grid(row=1, column=1, padx=10, pady=10)

    prev_button = tk.Button(window, text="이전", command=prev_words)
    prev_button.grid(row=1, column=0, padx=10, pady=10)
    
    window.mainloop()

# 엑셀 파일 경로
file_path = r"C:\Eng_dic\interface\OnlyCarrot\WordList.xlsx"

# 엑셀 파일 읽어오기
word_data = read_excel(file_path)

# 단어 조회 GUI 실행
show_words(word_data, start_index=0)
