import tkinter as tk
from tkinter import messagebox
from openpyxl import load_workbook

def delete_word_from_excel(eng_word, file_path):
    # Load the Excel workbook
    wb = load_workbook(file_path)
    
    # Select the worksheet named 'wordsheet'
    ws = wb['wordsheet']
    
    found = False
    
    # Find the row containing the word to delete
    for row_num, row in enumerate(ws.iter_rows(min_row=2, max_col=1, max_row=ws.max_row, values_only=True), start=2):
        if eng_word in row:
            found = True
            # Delete the row
            ws.delete_rows(row_num)
            # Save the changes to the Excel file
            wb.save(file_path)
            # Display a message indicating success
            messagebox.showinfo("Success", f"단어 '{eng_word}'가 삭제되었습니다.")
            break
    
    # If the word is not found, display an error message
    if not found:
        messagebox.showerror("Error", f"단어 '{eng_word}'를 찾을 수 없습니다.")



def delete_word_gui():
    # Create the tkinter window
    window = tk.Tk()
    window.title("Delete Word")
    
    # Label and Entry for English word
    eng_label = tk.Label(window, text="영어단어:")
    eng_label.grid(row=0, column=0, padx=10, pady=5, sticky="w")
    eng_entry = tk.Entry(window)
    eng_entry.grid(row=0, column=1, padx=10, pady=5)
    
    # Function to delete word from Excel
    def delete_word():
        eng_word = eng_entry.get()
        delete_word_from_excel(eng_word, r"C:\Eng_dic\interface\OnlyCarrot\WordList.xlsx")
        # Clear the entry field after deleting word
        eng_entry.delete(0, 'end')
    
    # Button to delete word
    delete_button = tk.Button(window, text="삭제하기", command=delete_word)
    delete_button.grid(row=1, columnspan=2, padx=10, pady=10)
    
    # Run the tkinter event loop
    window.mainloop()

# Example usage
delete_word_gui()
