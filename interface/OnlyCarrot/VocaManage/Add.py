import tkinter as tk
from tkinter import messagebox
from openpyxl import load_workbook

def is_duplicate(eng_word, file_path):
    # Load the Excel workbook
    wb = load_workbook(file_path)
    
    # Select the worksheet named 'wordsheet'
    ws = wb['wordsheet']
    
    # Iterate through each row in column A (English words)
    for row in ws.iter_rows(min_row=2, max_col=1, max_row=ws.max_row, values_only=True):
        if eng_word in row:
            return True
    
    return False

def add_word_to_excel(eng_word, kor_meaning, word_class, file_path):
    # Check if the word is already in the Excel file
    if is_duplicate(eng_word, file_path):
        messagebox.showerror("Error", "중복된 단어입니다.")
        return
    
    # Load the Excel workbook
    wb = load_workbook(file_path)
    
    # Select the worksheet named 'wordsheet'
    ws = wb['wordsheet']
    
    # Find the first empty row in the worksheet
    empty_row = ws.max_row + 1
    
    # Write the data to the next empty row
    ws[f'A{empty_row}'] = eng_word
    ws[f'B{empty_row}'] = kor_meaning
    ws[f'C{empty_row}'] = word_class
    
    # Save the changes to the Excel file
    wb.save(file_path)
    
    # Display a message indicating success
    messagebox.showinfo("Success", "Word added successfully!")

def add_word_gui():
    # Create the tkinter window
    window = tk.Tk()
    window.title("Add Word")
    
    # Label and Entry for English word
    eng_label = tk.Label(window, text="영어단어:")
    eng_label.grid(row=0, column=0, padx=10, pady=5, sticky="w")
    eng_entry = tk.Entry(window)
    eng_entry.grid(row=0, column=1, padx=10, pady=5)
    
    # Label and Entry for Korean meaning
    kor_label = tk.Label(window, text="한글:")
    kor_label.grid(row=1, column=0, padx=10, pady=5, sticky="w")
    kor_entry = tk.Entry(window)
    kor_entry.grid(row=1, column=1, padx=10, pady=5)
    
    # Label and Entry for word class
    class_label = tk.Label(window, text="품사:")
    class_label.grid(row=2, column=0, padx=10, pady=5, sticky="w")
    class_entry = tk.Entry(window)
    class_entry.grid(row=2, column=1, padx=10, pady=5)
    
    # Function to add word to Excel
    def add_word():
        eng_word = eng_entry.get()
        kor_meaning = kor_entry.get()
        word_class = class_entry.get()
        add_word_to_excel(eng_word, kor_meaning, word_class, r"C:\Eng_dic\interface\OnlyCarrot\WordList.xlsx")
        # Clear the entry fields after adding word
        eng_entry.delete(0, 'end')
        kor_entry.delete(0, 'end')
        class_entry.delete(0, 'end')
    
    # Button to add word
    add_button = tk.Button(window, text="추가하기", command=add_word)
    add_button.grid(row=3, columnspan=2, padx=10, pady=10)
    
    # Run the tkinter event loop
    window.mainloop()

# Example usage
add_word_gui()
